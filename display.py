import re
import openpyxl

# Give the location of the file
path = "C:\\Users\\alvar\\Desktop\\data.xlsx"

bytecode = {
  "00" : "STOP",
  "01" : "ADD",
  "02" : "MUL",
  "03" : "SUB",
  "04" : "DIV",
  "05" : "SDIV",
  "06" : "MOD",
  "07" : "SMOD",
  "08" : "ADDMOD",
  "09" : "MULMOD",
  "0a" : "EXP",
  "0b" : "SIGNEXTEND",
  "10" : "LT",
  "11" : "GT",
  "12" : "SLT",
  "13" : "SGT",
  "14" : "EQ",
  "15" : "ISZERO",
  "16" : "AND",
  "17" : "OR",
  "18" : "XOR",
  "19" : "NOT",
  "1a" : "BYTE",
  "20" : "SHA3",
  "30" : "ADDRESS",
  "31" : "BALANCE",
  "32" : "ORIGIN",
  "33" : "CALLER",
  "34" : "CALLVALUE",
  "35" : "CALLDATALOAD",
  "36" : "CALLDATASIZE",
  "37" : "CALLDATACOPY",
  "38" : "CODESIZE",
  "39" : "CODECOPY",
  "3a" : "GASPRICE",
  "3b" : "EXTCODESIZE",
  "3c" : "EXTCODECOPY",
  "40" : "BLOCKHASH",
  "41" : "COINBASE",
  "42" : "TIMESTAMP",
  "43" : "NUMBER",
  "44" : "DIFFICULTY",
  "45" : "GASLIMIT",
  "50" : "POP",
  "51" : "MLOAD",
  "52" : "MSTORE",
  "53" : "MSTORE8",
  "54" : "SLOAD",
  "55" : "SSTORE",
  "56" : "JUMP",
  "57" : "JUMPI",
  "58" : "PC",
  "59" : "MSIZE",
  "5a" : "GAS",
  "5b" : "JUMPDEST",
  "60" : "PUSH1",
  "61" : "PUSH2",
  "62" : "PUSH3",
  "63" : "PUSH4",
  "64" : "PUSH5",
  "65" : "PUSH6",
  "66" : "PUSH7",
  "67" : "PUSH8",
  "68" : "PUSH9",
  "69" : "PUSH10",
  "6a" : "PUSH11",
  "6b" : "PUSH12",
  "6c" : "PUSH13",
  "6d" : "PUSH14",
  "6e" : "PUSH15",
  "6f" : "PUSH16",
  "70" : "PUSH17",
  "71" : "PUSH18",
  "72" : "PUSH19",
  "73" : "PUSH20",
  "74" : "PUSH21",
  "75" : "PUSH22",
  "76" : "PUSH23",
  "77" : "PUSH24",
  "78" : "PUSH25",
  "79" : "PUSH26",
  "7a" : "PUSH27",
  "7b" : "PUSH28",
  "7c" : "PUSH29",
  "7d" : "PUSH30",
  "7e" : "PUSH31",
  "7f" : "PUSH32",
  "80" : "DUP1",
  "81" : "DUP2",
  "82" : "DUP3",
  "83" : "DUP4",
  "84" : "DUP5",
  "85" : "DUP6",
  "86" : "DUP7",
  "87" : "DUP8",
  "88" : "DUP9",
  "89" : "DUP10",
  "8a" : "DUP11",
  "8b" : "DUP12",
  "8c" : "DUP13",
  "8d" : "DUP14",
  "8e" : "DUP15",
  "8f" : "DUP16",
  "90" : "SWAP1",
  "91" : "SWAP2",
  "92" : "SWAP3",
  "93" : "SWAP4",
  "94" : "SWAP5",
  "95" : "SWAP6",
  "96" : "SWAP7",
  "97" : "SWAP8",
  "98" : "SWAP9",
  "99" : "SWAP10",
  "9a" : "SWAP11",
  "9b" : "SWAP12",
  "9c" : "SWAP13",
  "9d" : "SWAP14",
  "9e" : "SWAP15",
  "9f" : "SWAP16",
  "a0" : "LOG0",
  "a1" : "LOG1",
  "a2" : "LOG2",
  "a3" : "LOG3",
  "f0" : "CREATE",
  "f1" : "CALL",
  "f2" : "CALLCODE",
  "f3" : "RETURN",
  "ff" : "SUICIDE"
}

#This function divides the bytecode in pairs in order to get the opcode
def prepare(str):

  if str is None:
    return

  str = str.replace('""', '')
  str = str[2:]
  x = re.findall("(..?)", str)
  return x

#This function gets the divided bytecode and translate each pair in their correspondent opcode
def translate(code):
  dis = ""
  i = 0
  il = iter(code)

  for x in il:
    try:
      opcode = bytecode[x]
      if opcode[:4] == "PUSH":
        #The instruction PUSHX uses the X following bytes, so they are added to the PUSH instruction and not count as new opcodes
        length = int(opcode.replace("PUSH", ""))
        aux = slice(i + 1, i + length + 1)
        st = code[aux]
        str1 = ''.join(st)
        dis = dis + opcode + " 0x" + str1 + "\n"
        i = i + length

        for k in range(length):
          #Skip the X bytes corresponding to the PUSH instruction
          next(il)

      else:
        dis = dis + opcode + "\n"
      i = i + 1

    except:
      error = ' Unknown Opcode'
      dis = dis + x + error + "\n"
      continue

  return dis


def read(path):
  # workbook object is created
  wb_obj = openpyxl.load_workbook(path)

  sheet_obj = wb_obj.active
  m_row = sheet_obj.max_row

  # Loop will get all values of first column
  # First value in my excel is the name of the colum (bytecode column), so the loop will start at pos 2
  for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    str= cell_obj.value
    code = prepare(str)

    if code is not None:
      result = translate(code)
      print (result)

read(path)
